import sys
import os
import smtplib
import sqlite3
from email.mime.text import MIMEText
from openpyxl import Workbook

# db stuff
db = sqlite3.connect('grok.db')
cursor = db.cursor()
cursor.execute('create table chamados (categoria text, qtd_chamados integer)')
categorias = [('Categoria A', 10),
              ('Categoria B', 20),
              ('Categoria C', 30),
              ('Categoria D', 40)]
cursor.executemany('insert into chamados values (?, ?)', categorias)
db.commit()

# worksheet
file = sys.argv[1] + '.xlsx'
wb = Workbook()
ws = wb.active
ws['A1'] = 'Categoria'
ws['B1'] = 'Qtd Chamados'
for row in cursor.execute('select * from chamados'):
    ws.append(row)
wb.save(file)
db.close()
qtd_chamados = {r[0].value: r[1].value for r in ws.iter_rows(min_row=2)
                if r[1].value > 15}

# email and ssh
if qtd_chamados:
    msg_body = '''Foram encontradas {} categorias com um volume > 15 chamados.
    Por favor analisar os casos {}  '''.format(len(qtd_chamados) + 1,
                                               qtd_chamados)
    msg = MIMEText(msg_body)
    msg['Subject'] = 'Teste'
    msg['To'] = 'alguem@dxc.com'
    msg['From'] = 'eu@dxc.com'
    s = smtplib.SMTP('smtp.dxc.com')
    s.send_message(msg)
    s.quit()

os.system("scp %s usuario@server.dxc:/relatorios/" % file)
